from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
import os
from datetime import datetime
from json import dump
import shutil
import openpyxl
from csv import reader
import logging
from time import time
# import threading

class MyTableWidget(QWidget):
    def __init__(self, parent):
        super(QWidget, self).__init__(parent)
        self.initial_logger()
        self.initTabsWidget()
        self.init_clean_up_tab()
        self.init_move_files_tab()
        self.init_txt_to_xlsx_tab()
        self.init_config_file_tab()
        self.init_split_files_tab()

    def set_main_fodler(self, text:str) -> None:
        self.path1.setText(text)
        self.main_line1.setText(text)
        self.line_found.setText(text)
        self.line_split_review.setText(text)

    # region clean_up_tab

    def init_clean_up_tab(self) -> None:
        '''Функция инициализации первого таба.\n'''
        # Create elements
        self.btn1 = QPushButton('Обзор', self.clean_up_tab)
        self.btn1.clicked.connect(lambda : self.main_line1.setText(
                QFileDialog().getExistingDirectory(None,'Select directory')))

        self.btn2 = QPushButton('Выполнить', self.clean_up_tab)
        self.btn2.clicked.connect(self.function_node_clean_up_tab)
        self.main_line1 = QLineEdit(self.clean_up_tab)
        self.line_extension = QLineEdit('.exe',self.clean_up_tab)
        self.line_rename_file = QLineEdit('test.txt',self.clean_up_tab)
        self.line_rename_file.setToolTip('Только txt файлы!')
        self.line_extension.setEnabled(False)
        self.line_rename_file.setEnabled(False)

        self.check_box1 = QCheckBox('Удаление файлов', self.clean_up_tab)
        self.check_box2 = QCheckBox('Переименование файлов', self.clean_up_tab)
        self.check_box3 = QCheckBox('Удаление пустых папок', self.clean_up_tab)
        self.check_box1.clicked.connect(lambda state: self.line_extension.setEnabled(state))
        self.check_box2.clicked.connect(lambda state: self.line_rename_file.setEnabled(state))

        self.text = QTextEdit(self.clean_up_tab)
        self.text.setReadOnly(True)

        # Create grid
        grid = QGridLayout()
        grid.addWidget(self.main_line1, 0, 0,1,6)
        grid.addWidget(self.btn1, 0, 6)
        grid.addWidget(self.check_box1, 1,0)
        grid.addWidget(self.line_extension,1,1, alignment=Qt.AlignCenter)
        grid.addWidget(self.check_box2, 2,0)
        grid.addWidget(self.line_rename_file, 2,1, alignment=Qt.AlignCenter)
        grid.addWidget(self.check_box3, 3,0)
        grid.addWidget(self.text, 4,0,5,7)
        grid.addWidget(self.btn2, 10, 6)

        self.clean_up_tab.setLayout(grid)

    def function_node_clean_up_tab(self) -> None:
        '''Функция проверки на наличие и корректности введённых данных для вызова определённой операции над файлами.'''
        self.text.clear()
        if not os.path.exists(self.main_line1.text()):
            QMessageBox.warning(self, 'Ошибка', 'Такого пути не существует')
            return
        t = time()
        if self.check_box1.isChecked() and self.line_extension.text():
            self.Print(self.text, 'Удаление...')
            tmp = self.delete_files()
            self.Print(self.text, f'Удалено {tmp[0]} файлов ({self.convert_bytes(tmp[1])})')
        if self.check_box2.isChecked() and self.line_rename_file.text():
            self.Print(self.text,'Переименование...')
            self.rename_files()
            self.Print(self.text,'Файлы переименованы...')
        if self.check_box3.isChecked():
            self.Print(self.text,'Удаление пустых папок...')
            self.delete_empty_directories(self.main_line1.text())
            self.Print(self.text,'Папки удалены...')
        print(time() - t)

        if not (self.check_box1.isChecked() or self.check_box2.isChecked() or self.check_box3.isChecked()):
            QMessageBox.warning(self, 'Ошибка', 'Выберите действие')

        self.disabled_check_boxes()

    def disabled_check_boxes(self) -> None:
        '''Функция снятия выделения чек-боксов.\n
        Она вызывается при успешной выполенении операции.'''
        self.check_box1.setChecked(False)
        self.check_box2.setChecked(False)
        self.check_box3.setChecked(False)
        self.line_extension.setEnabled(False)
        self.line_rename_file.setEnabled(False)

    def convert_bytes(self,size) -> str:
        '''Функция определения единиц измерения по размеру удалённых файлов.\n
        Возрвращает единицы измерения и округлённый до целых размер удалённых файлов.'''
        for x in ['байтов', 'Кб', 'Мб', 'Гб', 'Тб']:
            if size < 1024.0:
                return "%3.1f %s" % (size, x)
            size /= 1024.0

        return size

    def delete_files(self) -> tuple[int,int]:
        '''Функция из главного функционала\n
        Одна из функции главного функционала программы.
        Удаляет файлы с определённым расширением или именем.'''
        size = 0
        removing_files = self.find_all_files_extension(
            self.main_line1.text(),
            self.line_extension.text())

        for i in removing_files:
            self.Print(self.text, f'Файл {i} удалён')
            size += os.stat(i).st_size
            try:
                os.remove(i)
            except Exception:
                logging.error('Exception', exc_info=True)
                return (removing_files.index(i),size)
        return (len(removing_files),size)

    def find_first_file(self, endswitch:str, files:list) -> str:
        '''Функция выбора первого файла из списка. Файлы ищутся с определённым расширением.'''
        for file in files:
            if file.endswith(endswitch):
                return file
        return ''

    def rename_files(self) -> None:
        '''Функция из главного функционала\n
        Выполняет поиск файлов с одинаковым именем (например Out_res.txt)
        и переименование по следующем конструкциям:\n
        1.\tназвание папки в которой находится файл + дата и время из найденого dat-файла (если есть хотя бы один в папке) + текущее имя;
        2.\tназвание папки в которой находится файл + текущее имя.'''
        for root, _, files in os.walk(self.main_line1.text()):
            dat =  self.find_first_file('.dat', files)
            file =  self.find_first_file(self.line_rename_file.text(), files)
            if not file: continue

            folder = root.split('/')[-1].split('\\')[-1]
            if not dat:
                new_name = '_'.join([folder,file])
            else:
                dat_split = '_'.join(dat.split('_')[1:5])
                new_name = '_'.join([folder,dat_split,file])

            try:
                os.rename(os.path.join(root,file),os.path.join(root,new_name))
                self.Print(self.text, f'{os.path.join(root,file)} -> {os.path.join(root,new_name)}')
            except Exception:
                logging.error('Exception', exc_info=True)

    def delete_empty_directories(self, path:str) -> None:
        '''Функция из главного функционала\n
        Выполняет удаление пустых директорий.'''
        for _dir in os.listdir(path):
            current_dir = os.path.join(path, _dir)
            if os.path.isdir(current_dir):
                self.delete_empty_directories(current_dir)
                if not os.listdir(current_dir):
                    try:
                        os.rmdir(current_dir)
                    except Exception as e:
                        print(e)
                    # self.catch_error(, self.text, current_dir)
                    self.Print(self.text,"Папка: " + current_dir + " удалена")


    # endregion

    # region move_files_tab

    def init_move_files_tab(self) -> None:
        '''Функция инициализации второго таба.'''
        # Create elements
        self.radio_button1 = QRadioButton('Сборка txt-файлов', self.move_files_tab)
        self.radio_button2 = QRadioButton('Рассылка исполняемых файлов', self.move_files_tab)
        self.radio_button1.setChecked(True)
        group = QButtonGroup(self.move_files_tab)
        group.addButton(self.radio_button1)
        group.addButton(self.radio_button2)

        self.label_1 = QLabel('Путь', self.move_files_tab)
        self.label_2 = QLabel('Путь', self.move_files_tab)

        self.path1 = QLineEdit(self.move_files_tab)
        self.path2 = QLineEdit(self.move_files_tab)
        self.area = QTextEdit(self.move_files_tab)
        self.area.setReadOnly(True)
        self.probar = QProgressBar(self.move_files_tab)
        self.probar.setAlignment(Qt.AlignCenter)

        self.btn1 = QPushButton('Обзор',self.move_files_tab)
        self.btn2 = QPushButton('Обзор',self.move_files_tab)
        self.btn1.clicked.connect(lambda : self.path1.setText(
                QFileDialog().getExistingDirectory(None,'Select directory')))
        self.btn2.clicked.connect(self.select_second_path)

        self.btn4 = QPushButton('Выполнить',self.move_files_tab)
        self.btn4.clicked.connect(self.function_node_move_files_tab)
        # Create grid for paint elements
        grid = QGridLayout()
        grid.setVerticalSpacing(30)
        grid.setRowStretch(6,2)
        grid.setColumnStretch(0,1)
        grid.setColumnStretch(1,5)
        grid.setColumnStretch(2,5)
        grid.setColumnStretch(3,5)


        grid.addWidget(self.radio_button1, 0,1, alignment=Qt.AlignCenter)
        grid.addWidget(self.radio_button2, 0,3, alignment=Qt.AlignLeft)

        grid.addWidget(self.label_1, 1,0)
        grid.addWidget(self.label_2, 2,0)

        grid.addWidget(self.path1, 1,1,1,3)
        grid.addWidget(self.path2, 2,1,1,3)
        grid.addWidget(self.area, 3,1,3,3)
        grid.addWidget(self.probar, 6, 1, 1, 3)

        grid.addWidget(self.btn1, 1, 4, 1,1)
        grid.addWidget(self.btn2, 2, 4, 1,1)
        grid.addWidget(self.btn4, 6, 4, 1,1, alignment=Qt.AlignBottom)

        self.move_files_tab.setLayout(grid)

    def select_second_path(self) -> None:
        '''Динамический выбор пути\n
        Учитывает выбор операции (радио кнопки).\n
        Первая операция (выбрано по умолчанию). Сборка txt-файлов требуется два пути:
        1.\tпуть поиска файлов (поиск выполняется и в подпапках по расширению txt)
        2.\tпуть куда они будут помещены\n
        Имеется проблема одинаковых имён файлов в месте хранения, в связи с чем необходимо их переименовывать в момент копирования.
        Новое имя определяется по шаблону "название папки + текущее имя файла"\n
        Вторая операция. Рассылка исполняемых файлов программы обработки, поэтому поля будут содержать следующие параметры:
        1.\tпуть куда куда нужно добавить исполняемый файл (добавляется и в подпапки)
        2.\tпуть к исполняемому файлу, который необходимо разослать.'''
        if self.radio_button1.isChecked():
            self.path2.setText(
                QFileDialog().getExistingDirectory(None,'Select directory'))
        else:
            self.path2.setText(
                QFileDialog.getOpenFileName(self,'Open File', None, '*.exe')[0])

    def function_node_move_files_tab(self) -> None:
        '''Функция проверки на наличие и корректности введённых параметров.'''

        if not self.path1.text():
            QMessageBox.warning(self, 'Error', 'First path is empty')

        if self.radio_button1.isChecked() and self.path2.text():
            self.Print(self.area, 'Копирование...')
            self.Print(self.area, f'Итого: {self.collect_files()} файлов скопировано')
            self.Print(self.area, 'Файлы скопированы.')


        if self.radio_button2.isChecked():
            self.Print(self.area, 'Начало рассылки...')
            self.send_out_files()
            self.Print(self.area, 'Файлы разосланы...')

    def collect_files(self) -> int:
        '''Функция из главного функционала.\n
        Выполняет сборку txt-файлов из подпапок в одну папку.
        Возвращает количество копированных файлов.'''
        files_list = self.find_all_files_extension(self.path1.text(),'.txt')
        for item in files_list:
            split = item.replace('\\','/').split('/')
            rename = f'{split[-2]}_{split[-1]}'
            to_path = os.path.join(self.path2.text(),rename)
            try:
                shutil.copyfile(item,to_path)
                percents = int((files_list.index(item)+1) / len(files_list)) * 100
                self.probar.setValue(percents)
                self.Print(self.area, f'{item} -> {to_path}')
            except Exception:
                logging.error('Exception', exc_info=True)
                return files_list.index(item) + 1
        return len(files_list)

    def send_out_files(self) -> None:
        '''Функция из главного функционала.\n
        Выполняет рассылку исполняемых файлов по папкам.'''
        exe = self.path2.text().split('/')[-1]
        for root,folders,_ in os.walk(self.path1.text()):
            for folder in folders:
                to = os.path.join(root, folder,exe)
                try:
                    shutil.copyfile(self.path2.text(),to)
                    self.Print(self.area, os.path.join(root, folder,exe))
                except Exception:
                    logging.error('Exception', exc_info=True)
                    return



    # endregion

    # region txt_to_xlsx_tab

    def init_txt_to_xlsx_tab(self) -> None:
        '''Функция инициализации третьего таба.'''
        self.line_found = QLineEdit(self.txt_to_xlsx_tab)
        self.line_name_one_to_one = QLineEdit(self.txt_to_xlsx_tab)
        self.line_name_one_to_one.setEnabled(False)
        self.line_name_many_to_one = QLineEdit(self.txt_to_xlsx_tab)
        self.line_name_many_to_one.setEnabled(False)
        self.line_save = QLineEdit(self.txt_to_xlsx_tab)

        self.btn_review_found = QPushButton('Обзор',self.txt_to_xlsx_tab)
        self.btn_review_found.clicked.connect(lambda : self.line_found.setText(
                QFileDialog().getExistingDirectory(None,'Select directory')))

        self.btn_review_save = QPushButton('Обзор',self.txt_to_xlsx_tab)
        self.btn_review_save.clicked.connect(lambda : self.line_save.setText(
                QFileDialog().getExistingDirectory(None,'Select directory')))

        self.btn_found = QPushButton('Найти',self.txt_to_xlsx_tab)
        self.btn_found.clicked.connect(self.found_files)
        self.btn_save = QPushButton('Выполнить',self.txt_to_xlsx_tab)
        self.btn_save.clicked.connect(self.function_node_txt_to_xlsx_tab)
        self.btn_hide = QPushButton('Снять выделения', self.txt_to_xlsx_tab)
        self.btn_hide.clicked.connect(self.hide_items)
        self.btn_select = QPushButton('Выделить всё', self.txt_to_xlsx_tab)
        self.btn_select.clicked.connect(self.select_items)

        self.label_select = QLabel(self.txt_to_xlsx_tab)
        self.label_select.setText('Выбрано: 0 файлов.')

        self.list = QListWidget(self.txt_to_xlsx_tab)
        self.list.setSelectionMode(QAbstractItemView.MultiSelection)



        self.checkbox_xlsx1 = QCheckBox('Каждый файл в отдельный лист',self.txt_to_xlsx_tab)
        self.checkbox_xlsx1.clicked.connect(lambda state: self.line_name_one_to_one.setEnabled(state))
        self.checkbox_xlsx2 = QCheckBox('Все файлы в один лист',self.txt_to_xlsx_tab)
        self.checkbox_xlsx2.clicked.connect(lambda state: self.line_name_many_to_one.setEnabled(state))

        grid = QGridLayout()
        grid.addWidget(self.line_found, 0,0,1,2)
        grid.addWidget(self.btn_review_found, 0,2)
        grid.addWidget(self.btn_found, 0,3)
        grid.addWidget(self.checkbox_xlsx1, 1,0)
        grid.addWidget(self.line_name_one_to_one, 1,1)
        grid.addWidget(self.checkbox_xlsx2, 2,0)
        grid.addWidget(self.line_name_many_to_one, 2,1)
        grid.addWidget(self.line_save, 3,0,1,2)
        grid.addWidget(self.btn_review_save, 3,2)

        grid.addWidget(self.btn_hide,4,3)
        grid.addWidget(self.btn_select,4,3, alignment=Qt.AlignTop)

        grid.addWidget(self.label_select, 5,0)

        grid.addWidget(self.list, 4,0,1,3)
        grid.addWidget(self.btn_save, 5,3)
        self.txt_to_xlsx_tab.setLayout(grid)

    def function_node_txt_to_xlsx_tab(self) -> None:
        '''Функция проверки на наличие и корректности введённых параметров.'''
        check1 = self.checkbox_xlsx1.isChecked()
        check2 = self.checkbox_xlsx2.isChecked()

        if not check1 and not check2:
            return QMessageBox.warning(self, 'Ошибка','Выберите вид сохранения эксель файла.')

        if not self.line_name_one_to_one.text() and not self.line_name_many_to_one.text():
            return QMessageBox.warning(self, 'Ошибка','Введите имя файла.')

        if len(self.list.selectedItems()) == 0:
            return QMessageBox.warning(self, 'Ошибка','Выберите файлы для сохранения.')

        if not self.line_save.text():
            return QMessageBox.warning(self, 'Ошибка','Выберите место для сохранения эксель документа.')


        if check1:
            self.one_file_to_one_sheet()
        if check2:
            self.many_files_to_one_sheet()

        QMessageBox.information(self,'Успех','Файлы переведены.')

    def delete_sheet(self,wb):
        '''Удаляет лист из эксель файла под именем "Sheet" (Он создаётся по умолчанию).'''
        del wb['Sheet']

    def many_files_to_one_sheet(self) -> None:
        '''Перебирает txt-файлы, считывая их по строке, и добавляет все файлы в один лист.'''
        _workbook = openpyxl.Workbook()
        _worksheet = _workbook.create_sheet('Excel')
        for item in self.list.selectedItems():
            with open(item.text(),'r') as file:
                cr = reader(file,delimiter = '\t')
                content = [line for line in cr]
                for line in content:
                    for i in range(1, len(line) - 1):
                        line[i] = line[i].replace(' ', '')
                        line[i] = float(line[i])
                    _worksheet.append(line)
                file.close()
        self.delete_sheet(_workbook)
        _workbook.save(filename=os.path.join(self.line_save.text(), self.line_name_many_to_one.text() + '.xlsx'))

    def one_file_to_one_sheet(self) -> None:
        '''Перебирает txt-файлы, считывая их по строке, и добавляет каждый файл в отдельный лист.'''
        _workbook = openpyxl.Workbook()
        for item in self.list.selectedItems():
            _worksheet = _workbook.create_sheet(item.text().split('\\')[-1])
            with open(item.text(),'r') as file:
                cr = reader(file,delimiter = '\t')
                content = [line for line in cr]


                for line in content:
                    for i in range(1, len(line) - 1):
                        line[i] = line[i].replace(' ', '')
                        line[i] = float(line[i])
                    _worksheet.append(line)
                file.close()
            _workbook.save(filename=os.path.join(self.line_save.text(),self.line_name_one_to_one.text() + '.xlsx'))
        self.delete_sheet(_workbook)

    def found_files(self) -> None:
        '''Функция поиска txt-файлов. Все текстовые документы (.txt) которые находит,
        добавляет в лист для выбора необходимых файлов перевода в эксель.'''
        self.list.clear()
        result = []
        if not self.line_found.text():
            return result
        self.list.addItems(self.find_all_files_extension(self.line_found.text(), '.txt'))
        self.list.itemClicked.connect(self.item_clicked)

    def changeItems(self, flag) -> None:
        '''Изменияет состояние элемента.'''
        for item in range(self.list.count()):
            self.list.item(item).setSelected(flag)
        self.item_clicked()

    def select_items(self) -> None:
        '''Выделяет все элементы'''
        return self.changeItems(True)

    def hide_items(self) -> None:
        '''Снимает все выделения'''
        return self.changeItems(False)

    def item_clicked(self) -> None:
        '''Изменяет лэйбл при выделении элемента. Показывает сколько элементов выбрано.'''
        return self.label_select.setText(f'Выбрано: {len(self.list.selectedItems())} элементов')

    # endregion

    # region config_file_tab

    def init_config_file_tab(self) -> None:
        '''Функция инициализации четвертого таба.\n'''
        # Create elements
        self.radio_config1 = QRadioButton('1200', self.config_file_tab)
        self.radio_config2 = QRadioButton('2500', self.config_file_tab)
        self.radio_config3 = QRadioButton('Local', self.config_file_tab)
        self.radio_config4 = QRadioButton('Reson', self.config_file_tab)
        self.radio_config5 = QRadioButton('2', self.config_file_tab)
        self.radio_config6 = QRadioButton('4', self.config_file_tab)

        self.group1 = QButtonGroup(self.config_file_tab)
        self.group1.addButton(self.radio_config1)
        self.group1.addButton(self.radio_config2)

        self.group2 = QButtonGroup(self.config_file_tab)
        self.group2.addButton(self.radio_config3)
        self.group2.addButton(self.radio_config4)

        self.group3 = QButtonGroup(self.config_file_tab)
        self.group3.addButton(self.radio_config5)
        self.group3.addButton(self.radio_config6)

        self.save_json_btn = QPushButton('Сохранить',self.config_file_tab)
        self.save_json_btn.clicked.connect(self.generate_config_file)

        grid = QGridLayout()
        grid.addWidget(QLabel('Радар'), 0,0)
        grid.addWidget(self.radio_config1, 0,1)
        grid.addWidget(self.radio_config2, 0,2)
        grid.addWidget(QLabel('Режим'), 1,0)
        grid.addWidget(self.radio_config3, 1,1)
        grid.addWidget(self.radio_config4, 1,2)
        grid.addWidget(QLabel('Канал'), 2,0)
        grid.addWidget(self.radio_config5, 2,1)
        grid.addWidget(self.radio_config6, 2,2)
        grid.addWidget(self.save_json_btn, 3,2, alignment=Qt.AlignCenter)
        grid.setRowStretch(4,2)
        grid.setVerticalSpacing(20)

        self.config_file_tab.setLayout(grid)

    def check_selected_buttons(self) -> bool:
        if self.group1.checkedButton() == None:
            QMessageBox.warning(self,'Ошибка','Выберите радар.')
            return False

        if self.group2.checkedButton() == None:
            QMessageBox.warning(self,'Ошибка','Выберите режим.')
            return False

        if self.group3.checkedButton() == None:
            QMessageBox.warning(self,'Ошибка','Выберите канал.')
            return False

        return True

    def generate_config_file(self) -> None:
        if not self.check_selected_buttons():return
        '''Генерирует файл конфигурации для программы обработки'''
        Radar = self.group1.checkedButton()     #Радар
        Mode = self.group2.checkedButton()      #Режим
        Channel = self.group3.checkedButton()   #Канал

        if Radar == None or Mode == None or Channel == None:
            return
        _dict = {
            'Radar':Radar.text(),
            'Mode':Mode.text(),
            'Channel':Channel.text()
        }
        try:
            with open('config.json', 'w') as json:
                dump(_dict,json,indent=4)
        except Exception:
            logging.error('Exception', exc_info=True)
            return
        logging.info('Конфиг файл сохранён')
        logging.info(_dict)
        QMessageBox.information(self, 'Успех', 'Файл успешно сохранён.')

    # endregion

    # region split_files_tab

    def init_split_files_tab(self) -> None:
        '''Функция инициализации пятого таба.\n'''
        self.radio_optimization = QRadioButton('Разделить файлы для оптимизации обработки', self.split_files_tab)
        self.radio_profile = QRadioButton('Разделение по профилям')

        self.radio_optimization.setChecked(True)

        self.line_split_review = QLineEdit(self.split_files_tab)
        self.line_first_file = QLineEdit(self.split_files_tab)
        self.line_first_file.setToolTip('Первый файл профиля')
        self.line_last_file = QLineEdit(self.split_files_tab)
        self.line_last_file.setToolTip('Последний файл профиля')
        self.line_profile_name = QLineEdit(self.split_files_tab)
        self.line_profile_name.setToolTip('Имя профиля')

        self.area_split = QTextEdit(self.split_files_tab)
        self.area_split.setReadOnly(True)

        self.radio_optimization.clicked.connect(self.optimizate_radio_elements)
        self.radio_profile.clicked.connect(self.profile_radio_elements)

        self.btn_split_review = QPushButton('Обзор', self.split_files_tab)
        self.btn_split_review.clicked.connect(lambda : self.line_split_review.setText(
                QFileDialog().getExistingDirectory(None,'Select directory')))

        self.btn_split = QPushButton('Выполнить', self.split_files_tab)
        self.btn_split.clicked.connect(self.function_node_split_files)
        self.combo = QComboBox(self.split_files_tab)
        self.combo.addItems(['2','3','4','5','6','7','8','9','10'])

        self.optimizate_radio_elements()
        group = QButtonGroup(self.split_files_tab)
        group.addButton(self.radio_optimization)
        group.addButton(self.radio_profile)


        grid = QGridLayout()
        grid.addWidget(QLabel('Путь'), 0,0)
        grid.addWidget(self.line_split_review, 0,1,1,4)
        grid.addWidget(self.btn_split_review, 0,5)
        grid.addWidget(self.radio_optimization, 1,0)
        grid.addWidget(self.combo, 1,1, alignment=Qt.AlignLeft)
        grid.addWidget(self.radio_profile, 2,0)
        grid.addWidget(self.line_first_file, 2,1)
        grid.addWidget(self.line_last_file, 2,2)
        grid.addWidget(self.line_profile_name, 2,3)
        grid.addWidget(self.area_split, 3,0,3,6)
        grid.addWidget(self.btn_split, 7,5)
        self.split_files_tab.setLayout(grid)

    def optimizate_radio_elements(self) -> None:
        '''Включает комбо-бокс и отключает три линии редактирования.'''
        self.combo.setDisabled(False)
        self.line_first_file.setDisabled(True)
        self.line_last_file.setDisabled(True)
        self.line_profile_name.setDisabled(True)

    def profile_radio_elements(self) -> None:
        '''Отключает комбо-бокс и включает три линии редактирования.'''
        self.combo.setDisabled(True)
        self.line_first_file.setDisabled(False)
        self.line_last_file.setDisabled(False)
        self.line_profile_name.setDisabled(False)

    def function_node_split_files(self) -> None:
        '''Функция проверки на наличие и корректности введённых параметров.'''
        if self.radio_optimization.isChecked():
            self.Print(self.area_split, f'Разделение файлов...')
            self.optimization()

        if self.radio_profile.isChecked() and self.create_profile_folder():
            self.Print(self.area_split, f'Разделение файлов...')
            self.copy_files(self.profiles())

        self.Print(self.area_split, f'Файлы разделены.')

    def get_files(self) -> list[str]:
        '''Возвращает список dat-файлов найденных в указанной директории.'''
        cwd = self.line_split_review.text()
        return [os.path.join(cwd,item) for item in os.listdir(cwd) if item.endswith('.dat')]

    def copy_files(self, files) -> None:
        '''Копирует список файлов в указанную папку'''
        for item in files:
            try:
                shutil.copyfile(item,os.path.join(self.profile_folder,item.split('\\')[-1]))
            except Exception:
                logging.info('Exception', exc_info=True)

    def create_profile_folder(self) -> bool:
        '''Создает папку для профиля и содержит список файлов по профилю.\n
        Возвращает:\n
        1.\tTrue  - папка успешно создана;
        2.\tFalse - папка с таким именем уже существует.'''
        self.profile_folder = os.path.join(self.line_split_review.text(), self.line_profile_name.text())
        if os.path.isdir(self.profile_folder):
            QMessageBox.warning(self, 'Ошибка', 'Папка уже существует')
            return False
        os.makedirs(self.profile_folder)
        return True

    def profiles(self) -> list:
        '''Возвращает список файлов, которые входят в профиль.'''
        files = self.get_files()
        res = []
        for item in range(len(files)):
            if files[item].endswith(self.line_first_file.text() + '.dat'):
                i = item
                while(not files[i].endswith(self.line_last_file.text() + '.dat')):
                    res.append(files[i])
                    i += 1
                res.append(files[i])

        return res

    def optimization(self) -> None:
        '''Разбивает файлы по количеству(QComboBox) указанных папок.'''
        split_num = int(self.combo.currentText())
        self.split_num = split_num
        folders = self.make_directories()
        dat_files = self.get_files()
        # поиск файлов dat
        count_dat_files = len(dat_files)
        split_list = [count_dat_files // split_num] * split_num
        split_list[-1] += count_dat_files % split_num
        for i in range(split_num):
            for _ in range(split_list[i]):
                try:
                    shutil.move(dat_files[0],os.path.join(folders[i],dat_files[0].replace('\\','/').split('/')[-1]))
                    dat_files.pop(0)
                except FileNotFoundError:
                    logging.error('Exception',exc_info=True)
                    return

    def make_directories(self) -> list[str]:
        '''Возрващает список созданных директорий.'''
        cwd = self.line_split_review.text()
        folders = []
        for i in range(self.split_num):
            try:
                folder = os.path.join(cwd,f'Часть_{i+1}')
                folders.append(folder)
                os.makedirs(folder)
            except FileExistsError:
                logging.error('Exception', exc_info=True)
                self.Print(self.area_split, f'Папка {folder} уже существует.')
                continue

        self.Print(self.area_split,f'Папки успешно созданы.')
        return folders

    # endregion

    def initTabsWidget(self) -> None:
        '''
        Функция инициализации табов
        '''
        self.tabs = QTabWidget()
        self.tabs.setTabPosition(QTabWidget.West)
        self.clean_up_tab = QWidget()
        self.move_files_tab = QWidget()
        self.txt_to_xlsx_tab = QWidget()
        self.config_file_tab = QWidget()
        self.split_files_tab = QWidget()
        self.tabs.resize(300,200)

        self.tabs.addTab(self.clean_up_tab,"Чистка")
        self.tabs.addTab(self.move_files_tab,"Перемещение файлов")
        self.tabs.addTab(self.txt_to_xlsx_tab,"Перевод txt to xlsx")
        self.tabs.addTab(self.config_file_tab,"Конфиг файл")
        self.tabs.addTab(self.split_files_tab,"Разделение файлов")

        vbox = QVBoxLayout()
        vbox.addWidget(self.tabs)
        self.setLayout(vbox)

    def find_all_files_extension(self, cwd:str, extension:str) -> list[str]:
        '''Функция поиска файлов по расширению.\n
        Возрващает список строк.\n
        Одна строка - это полный путь к файлу.
        '''
        files_list = []
        for root, _, files in os.walk(cwd):
            for item in files:
                if item.endswith(extension):
                    files_list.append(os.path.join(root,item))
        return files_list

    def Print(self,textEdit,string:str) -> None:
        '''Функция печати результата.\n
        Принимает объект, в котором будет отображена строка, которая передаётся третьим параметром.'''
        logging.info(string)
        textEdit.append(f'[{datetime.now().strftime("%d-%m-%Y %H:%M:%S")}] {string}')

    def initial_logger(self):
        logging.basicConfig(filename='pupo.log',format='%(asctime)s - %(levelname)s - %(message)s', level = logging.INFO)
        logging.info('Программа запущена')
